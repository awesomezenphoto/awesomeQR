'''
Function: 
0. command will take two input files, 
ex. awe-att.py <class-records> <attendants>
1. Find the first, last name in xlsx file, 
Note: openpyxl only support xlsx, NOT xls.... save the excl file to xlsx format
2. Read the Date string from attendants file, the first line must be the date string
	and the student names.
Note: ex.  2019-12-07, which is the excel date format reading from python.
3. Find the attendant name in class-records file, and update the check-in 
4. Save to OutFile .xlsx 
'''
import openpyxl, datetime, sys, os, string

Revision = "0.1"
FIRSTNAME = "First Name"
LASTNAME  = "Last Name"
#DATEAT = "2019-12-07"
OutFile = 'awe-upd.xlsx'

def find_value_in_two_column(ws, fname, lname, FirstNameCol, LastNameCol ):
	row = 1
	while True:
		row +=1 
		fn = FirstNameCol+str(row)
		ln = LastNameCol+str(row)
		#print (fn, ln)
		fcell = ws[fn]
		lcell = ws[ln]
		if fcell.value == None or lcell.value == None : 
			return None
		name = fcell.value.strip() + " " + lcell.value.strip()
		#print ("cell", name.strip(), fname+lname)
		if ( name.strip() == fname+ " " + lname ) :
			#print ( "Found it at ", row)
			return row 
	return None 		
				
	
def find_value_in_column(ws, search_string, column):
	row = 1
	for cell in ws[column]:
		#print (cell.value)
		if search_string in str(cell.value):
			return column, row
		row += 1
	return column, None 
			
			
def main():
	print("AWESOME Check-in Update python scrip Rev:", Revision)
	argc = len(sys.argv) - 1
	if ( argc < 2 ):
		print("Missing Input File")
		print("ex. awe-att.exe <class.xlsx> <attendant>")
		print("addendant file must start with date string yyyy-mm-dd")
		sys.exit()
	print (argc, sys.argv[1], sys.argv[2])
	try:
		book = openpyxl.load_workbook(sys.argv[1])
	except:
		sys.exit("ERROR. File is NOT Found from Your Input FileName, Please Try Again...", sys.argv[1])
	try:
		attd = open(sys.argv[2], "r")
	except:
		sys.exit("ERROR. File is NOT Found from Your Input FileName, Please Try Again...", sys.argv[1])
	# the first line must be the date of the class
	DateStr = attd.readline()
	DateStr = DateStr.strip()
	# read all lines from file
	students = attd.readlines()
	attd.close()
	sheet = book.active
	#b2 = sheet['B2']
	#c2 = sheet['c2']
	#print (b2.value, c2.value)
	FirstNameCol = ""
	LastNameCol = ""
	DateCol =""
	for col in ["A","B","C","D","E"]:
		frtcol, frtrow = find_value_in_column(sheet, FIRSTNAME, col) 
		if frtrow != None :
			print ("Found:", FIRSTNAME, "at", frtcol, frtrow)
			FirstNameCol = col
			break

	for col in ["A","B","C","D","E"]:
		lrtcol, lrtrow = find_value_in_column(sheet, LASTNAME, col) 
		if lrtrow != None :
			print ("Found:", LASTNAME, "at", lrtcol, lrtrow)
			LastNameCol = col
			break

	
	# Date fomat in xlsx file 2019-09-14 00:00:00
	for i in range(ord('A'), ord('Z')+1):
		drtcol, drtrow = find_value_in_column(sheet, DateStr, chr(i)) 
		if drtrow != None :
			print ("Found:", DateStr, "at", drtcol, drtrow)
			DateCol = chr(i)
			break
		
	
	print (FIRSTNAME,"column:",FirstNameCol,";", LASTNAME,"column:", LastNameCol,";", "DateStr column:", DateCol )
	if (FirstNameCol=="" or LastNameCol=="" or DateCol=="" ):
		if (FirstNameCol=="" or LastNameCol=="" ):
			print ("Missing", FIRSTNAME, "or", LASTNAME, "in the xlsx File")
		if (DateCol=="" ):
			print ("Date Str (column) are NOT matched between xlsx and attendant file!" )
		print ("Please Fix it and try again")
		sys.exit()
	stcount = 0
	for st in students:
		name=" ".join(st.split())
		fname , lname = name.split(' ', 1)
		row = find_value_in_two_column(sheet, fname, lname, FirstNameCol, LastNameCol )
		if ( row == None ):
			print ("Attendant:", fname, lname, "is NOT Registed.????????????????")
		else:
			d=DateCol+str(row)
			c = sheet[d]
			#d.alignment = Alignment(horizontal='center')
			if ( c.value == None ):
				sheet[d] = "x"
				print ("Attendant:", fname, lname, "is check-in")
				stcount += 1
			else:
				print ("Attendant:", fname, lname, "has been check-in ......")
				
	book.save(OutFile)	
	print ("====================================")
	print ("updated:", stcount, "records")
	print ("Output File is Saved as", OutFile)
  
if __name__== "__main__":
	main()			